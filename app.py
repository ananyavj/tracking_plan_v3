import streamlit as st
import streamlit.components.v1 as components
import os
import json
from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv()  # loads variables from .env into os.environ
from gemini_agent import run_gemini_audit_agent

#app.py 

st.set_page_config(
    page_title="Tracking Plan Auditor (Claude MCP)",
    page_icon="🤖",
    layout="wide"
)

# ─── HELPERS ─────────────────────────────────────────────────────────────────
def parse_tracking_plan(uploaded_file):
    # Parse the Excel tracking plan into a structured dict
    wb = load_workbook(uploaded_file, read_only=True)
    plan = {}
    sheets_to_parse = ['Identify', 'Page', 'Browsing', 'Purchase Funnel', 'Post-Purchase', 'Marketing']
    
    for sheet_name in sheets_to_parse:
        if sheet_name not in wb.sheetnames: continue
        ws = wb[sheet_name]
        current_event = None
        
        for row in ws.iter_rows(values_only=True):
            if not any(row): continue
            cell0 = str(row[0]).strip() if row[0] else ""
            cell1 = str(row[1]).strip() if row[1] else ""
            
            if cell0 in ("Event Name", "Property", "") or cell0.isupper() or cell0.startswith("▶"): continue
            
            if cell0 and cell1 and len(row) >= 4:
                req = str(row[2]).strip() if row[2] else "Optional"
                typ = str(row[3]).strip() if row[3] else "string"
                allowed = str(row[5]).strip() if len(row) > 5 and row[5] else None
                
                if cell0 not in plan:
                    plan[cell0] = {"properties": {}}
                    current_event = cell0
                    
                plan[cell0]["properties"][cell1] = {
                    "required": req == "Required",
                    "type": typ.lower(),
                    "allowed_values": [v.strip() for v in allowed.split("|")] if allowed else None
                }
    return plan

# ─── SIDEBAR ─────────────────────────────────────────────────────────────────
with st.sidebar:
    st.title("🤖 MCP Audit Config")
    st.markdown("---")
    
    st.subheader("1. Tracking plan")
    uploaded_plan = st.file_uploader("Upload Excel tracking plan", type=["xlsx"])
    
    st.subheader("2. AI Intelligence Provider")
    ai_provider = st.radio("Select AI Provider", ["Gemini"])
    
    if ai_provider == "Gemini":
        google_api_key = st.text_input("Google AI Studio API Key", type="password",
            value=os.getenv("GOOGLE_API_KEY", ""),
            help="Loaded from .env — paste a new key here to override.")
        anthropic_api_key = None
    
    st.subheader("3. Data Source (Amplitude)")
    amplitude_api_key = st.text_input("Amplitude API Key", type="password",
        value=os.getenv("AMPLITUDE_API_KEY", ""))
    amplitude_secret = st.text_input("Amplitude Secret Key", type="password",
        value=os.getenv("AMPLITUDE_SECRET_KEY", ""))
    amplitude_project_id = st.text_input("Amplitude Project ID", type="password",
        value=os.getenv("AMPLITUDE_PROJECT_ID", ""))
    eu_datacenter = st.checkbox("EU datacenter", value=False)
    
    st.markdown("---")
    run_btn = st.button("▶ Run MCP Audit", type="primary", use_container_width=True)
    
    st.markdown("---")
    st.caption("Fallback: Upload JSON to mock Amplitude API returns (for local testing without auth)")
    local_file = st.file_uploader("Local events JSON", type=["json"])

# ─── MAIN ────────────────────────────────────────────────────────────────────
st.title("🤖 Agentic Tracking Plan Auditor")
st.caption("AI connects directly to Amplitude via functional MCP to autonomously fetch and audit live tracking data.")

if not uploaded_plan:
    st.info("Upload your Excel tracking plan in the sidebar to get started.")
    st.stop()

plan = {}
with st.spinner("Parsing tracking plan..."):
    plan = parse_tracking_plan(uploaded_plan)
st.success(f"Tracking plan loaded — {len(plan)} events defined")

if run_btn:
    if ai_provider == "Gemini" and not google_api_key:
        st.error("Missing Google AI Studio API Key! This is required to run the Gemini Audit Agent.")
        st.stop()
        
    fallback_events = None
    if local_file:
        fallback_events = json.load(local_file)
        if isinstance(fallback_events, dict) and "events" in fallback_events:
            fallback_events = fallback_events["events"]
            
    if not fallback_events and not amplitude_api_key:
        st.error("Please provide Amplitude credentials to fetch real data, or upload a fallback local JSON.")
        st.stop()

    # Load system prompt
    system_prompt = ""
    try:
        with open("SKILL.md", "r", encoding="utf-8") as f:
            system_prompt = f.read()
    except FileNotFoundError:
        st.error("Could not find SKILL.md for the Agent prompt.")
        st.stop()

    app_config = {
        "api_key": amplitude_api_key,
        "secret_key": amplitude_secret,
        "project_id": amplitude_project_id,
        "eu_datacenter": eu_datacenter,
        "fallback_events": fallback_events
    }
    
    status_msg = st.empty()
    def update_status(msg):
        status_msg.info(f"🔄 {msg}")

    with st.spinner("Agent workflow activated..."):
        if ai_provider == "Gemini":
            report = run_gemini_audit_agent(
                google_api_key=google_api_key,
                system_prompt=system_prompt,
                tracking_plan=plan,
                app_config=app_config,
                status_callback=update_status
            )
    
    # Clear the status indicator text
    status_msg.empty()

    if "error" in report:
        st.error(f"Agent Execution Error: {report['error']}")
        with st.expander("Raw Output"):
            st.write(report.get("raw", ""))
    elif report is None:
        st.error(f"No JSON report was retrieved from the selected Agent.")
    else:
        st.success(f"{ai_provider} Audit Complete!")
        
        # Display the HTML report from JSON output (SKILL.md standard layout)
        html_report = report.get("html_report")
        if html_report:
            components.html(html_report, height=2000, scrolling=True)
        else:
            st.warning("No html_report found in the response. Showing raw JSON.")
            st.json(report)
            
        # Download raw findings JSON
        st.download_button(
            f"⬇ Download raw {ai_provider} findings (JSON)",
            data=json.dumps(report, indent=2),
            file_name=f"{ai_provider.lower()}_audit_findings.json",
            mime="application/json"
        )
