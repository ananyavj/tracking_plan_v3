# tracking_plan_parser_v2.py
"""
Kaliper V2 High-Resolution Parser

Interprets the Excel tracking plan into a deterministic rule map.
Handles dot-notation (nested paths) and platform-specific overrides.

PRINCIPLE: Preserve the raw complexity of the Excel. No simplified abstractions.
"""

import openpyxl
import json
import os
from datetime import datetime

EVENT_SHEETS = [
    "Identify", "Page", "Browsing", "Purchase Funnel", 
    "Post-Purchase", "Marketing"
]

# Assuming structure from existing parser
COL_EVENT_NAME  = 0
COL_PROPERTY    = 1 # Dot-notation expected here
COL_REQ_OPT     = 2
COL_TYPE        = 3
COL_EXAMPLE     = 4
COL_ALLOWED     = 5
COL_DESCRIPTION = 6
COL_NOTES       = 7 
COL_PLATFORM    = 8 # New assumption from Execution Lock

TYPE_NORMALIZE = {
    "string": "string", "str": "string",
    "float": "float", "number": "float", "decimal": "float",
    "integer": "integer", "int": "integer",
    "boolean": "boolean", "bool": "boolean",
    "array": "array", "list": "array",
    "object": "object", "dict": "object"
}

class TrackingPlanParserV2:
    @staticmethod
    def parse(filepath):
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"Tracking plan not found: {filepath}")
        
        wb = openpyxl.load_workbook(filepath, data_only=True)
        schema = {
            "events": {},
            "global_props": [],
            "meta": {
                "source": str(filepath),
                "parsed_at": datetime.now().isoformat()
            }
        }

        # 1. Parse Event Sheets
        for sheet_name in EVENT_SHEETS:
            if sheet_name not in wb.sheetnames: continue
            ws = wb[sheet_name]
            current_event = None

            for row_cells in ws.iter_rows(min_row=3, values_only=True):
                row = list(row_cells)
                evt_name_cell = TrackingPlanParserV2._get_cell(row, COL_EVENT_NAME)
                prop_cell     = TrackingPlanParserV2._get_cell(row, COL_PROPERTY)
                
                # Detective Work: Header vs Property
                is_header = False
                if evt_name_cell and ("▶" in evt_name_cell or not prop_cell):
                    is_header = True
                
                if is_header and evt_name_cell:
                    name = evt_name_cell.replace("▶", "").strip()
                    if not name: continue
                    current_event = name
                    schema["events"][current_event] = {
                        "sheet": sheet_name,
                        "rules": {
                            "global": [],
                            "ios": [],
                            "android": [],
                            "web": []
                        }
                    }
                    continue

                if prop_cell and current_event:
                    rule = TrackingPlanParserV2._build_rule(row)
                    platform = TrackingPlanParserV2._get_cell(row, COL_PLATFORM) or "all"
                    platform = platform.lower().strip()

                    if platform in ("all", "blank", ""):
                        schema["events"][current_event]["rules"]["global"].append(rule)
                    elif platform in schema["events"][current_event]["rules"]:
                        schema["events"][current_event]["rules"][platform].append(rule)
                    else:
                        # Default to global if platform is unknown but not all
                        schema["events"][current_event]["rules"]["global"].append(rule)

        # 2. Parse Global Props
        if "Global Props" in wb.sheetnames:
            ws = wb["Global Props"]
            for row_cells in ws.iter_rows(min_row=3, values_only=True):
                row = list(row_cells)
                prop = TrackingPlanParserV2._get_cell(row, 0)
                if prop and not prop.startswith("─"):
                    schema["global_props"].append(TrackingPlanParserV2._build_rule(row))

        return schema

    @staticmethod
    def _build_rule(row):
        allowed_raw = TrackingPlanParserV2._get_cell(row, COL_ALLOWED)
        allowed = [v.strip() for v in allowed_raw.split("|") if v.strip()] if allowed_raw else []
        
        req_raw = TrackingPlanParserV2._get_cell(row, COL_REQ_OPT)
        is_req = req_raw and "required" in req_raw.lower()
        
        return {
            "name":      TrackingPlanParserV2._get_cell(row, COL_PROPERTY).strip(),
            "required":  is_req,
            "type":      TYPE_NORMALIZE.get((TrackingPlanParserV2._get_cell(row, COL_TYPE) or "").lower(), "string"),
            "allowed":   allowed,
            "condition": TrackingPlanParserV2._get_cell(row, COL_NOTES) or ""
        }

    @staticmethod
    def _get_cell(row, idx):
        try:
            val = row[idx]
            return str(val).strip() if val is not None else None
        except: return None

if __name__ == "__main__":
    parser = TrackingPlanParserV2()
    res = parser.parse("tracking_plan.xlsx")
    print(json.dumps(res, indent=2))
