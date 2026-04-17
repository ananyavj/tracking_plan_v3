# tracking_plan_parser.py
"""
Parses the Fashion Marketplace Tracking Plan Excel into a JSON schema
consumed by the audit engine and AI agents.

Phase 3 change: Column H (implementation notes) is now captured as a
`condition` field on each property. The audit engine evaluates this string
before flagging M1/M2, allowing rules like "Populate on step 3 only."

FIX — Property row detection: The original condition required BOTH the
Event column (col A) AND the Property column (col B) to be non-empty for a
property row. In the actual Excel, col A is BLANK on property rows (only
filled on the event header row). The fix: a row is a property row if
col B (property name) is non-empty AND we already have a current_event in
scope, regardless of col A.
"""

import openpyxl
import json
from pathlib import Path

EVENT_SHEETS = [
    "Identify",
    "Page",
    "Browsing",
    "Purchase Funnel",
    "Post-Purchase",
    "Marketing",
]

# Column indices (0-based) in event sheets
COL_EVENT_NAME  = 0
COL_PROPERTY    = 1
COL_REQ_OPT     = 2
COL_TYPE        = 3
COL_EXAMPLE     = 4
COL_ALLOWED     = 5
COL_DESCRIPTION = 6
COL_NOTES       = 7   # Phase 3: captured as `condition`

# Column indices in Global Props sheet
GCOL_PROPERTY   = 0
GCOL_REQ_OPT    = 1
GCOL_TYPE       = 2
GCOL_EXAMPLE    = 3
GCOL_AMP_MAP    = 4
GCOL_DESC       = 5

TYPE_NORMALIZE = {
    "string":    "string",
    "str":       "string",
    "float":     "float",
    "number":    "float",
    "decimal":   "float",
    "integer":   "integer",
    "int":       "integer",
    "boolean":   "boolean",
    "bool":      "boolean",
    "array":     "array",
    "list":      "array",
    "iso8601":   "ISO8601",
    "datetime":  "ISO8601",
    "timestamp": "ISO8601",
}


def _cell(row, idx):
    try:
        val = row[idx]
        if val is None:
            return None
        s = str(val).strip()
        return s if s else None
    except IndexError:
        return None


def _normalize_type(raw):
    if raw is None:
        return "string"
    return TYPE_NORMALIZE.get(raw.lower().strip(), "string")


def _parse_allowed(raw):
    if not raw:
        return []
    return [v.strip() for v in raw.split("|") if v.strip()]


def _is_required(raw):
    if raw is None:
        return False
    return "required" in raw.lower()


def parse_event_sheets(wb):
    events = []

    for sheet_name in EVENT_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        current_event = None

        for row in ws.iter_rows(min_row=3, values_only=True):
            row = list(row)
            evt_col = _cell(row, COL_EVENT_NAME)
            prop    = _cell(row, COL_PROPERTY)

            # --- 1. DETECT EVENT HEADER ---
            # Primary signal: "▶" prefix (unambiguous, never appears on property rows).
            # Secondary fallback: evt_col is non-empty AND prop is blank AND evt_col
            # doesn't look like a property name (no spaces following a colon, not all-lowercase).
            # This prevents merged-cell spill from resetting current_event mid-sheet.
            is_event_header = False
            if evt_col:
                if "▶" in evt_col:
                    is_event_header = True
                elif not prop:
                    # Only treat as a header if it looks like an event name
                    # (title-cased words, not a property value like "required" or a number).
                    looks_like_event = (
                        len(evt_col) > 2
                        and not evt_col.replace(" ", "").isdigit()
                        and evt_col[0].isupper()
                    )
                    if looks_like_event:
                        is_event_header = True

            if is_event_header:
                event_name = evt_col.replace("▶", "").strip()
                if not event_name:
                    continue
                current_event = {
                    "event_name":  event_name,
                    "sheet":       sheet_name,
                    "description": _cell(row, COL_DESCRIPTION) or _cell(row, COL_NOTES) or event_name,
                    "properties":  [],
                }
                events.append(current_event)
                continue

            # --- 2. DETECT PROPERTY ROW ---
            # FIX: Only require `prop` to be non-empty (col A is blank on property rows).
            # The old condition `if prop and current_event and evt_col` never matched
            # because evt_col is always blank on property rows in the actual Excel.
            if prop and current_event:
                notes = _cell(row, COL_NOTES) or ""
                desc  = _cell(row, COL_DESCRIPTION) or ""

                # Logic: capture step-specific conditions from Notes or Description
                condition = notes if ("step" in notes.lower()) else desc if ("step" in desc.lower()) else ""

                current_event["properties"].append({
                    "name":           prop.strip(),
                    "required":       _is_required(_cell(row, COL_REQ_OPT)),
                    "type":           _normalize_type(_cell(row, COL_TYPE)),
                    "example":        _cell(row, COL_EXAMPLE) or "",
                    "allowed_values": _parse_allowed(_cell(row, COL_ALLOWED)),
                    "description":    desc,
                    "condition":      condition,
                })
                continue

            # --- 3. IGNORE BLANK ROWS ---
            # Prevents context reset for orphan properties below blank lines
            if not evt_col and not prop:
                continue

    return events


def parse_global_props(wb):
    if "Global Props" not in wb.sheetnames:
        return []

    ws = wb["Global Props"]
    props = []

    for row in ws.iter_rows(min_row=3, values_only=True):
        row  = list(row)
        prop = _cell(row, GCOL_PROPERTY)
        if not prop or prop.startswith("─"):
            continue
        props.append({
            "name":          prop,
            "required":      _is_required(_cell(row, GCOL_REQ_OPT)),
            "type":          _normalize_type(_cell(row, GCOL_TYPE)),
            "example":       _cell(row, GCOL_EXAMPLE) or "",
            "amplitude_map": _cell(row, GCOL_AMP_MAP) or "",
            "description":   _cell(row, GCOL_DESC) or "",
            "condition":     "",
        })

    return props


def parse_data_dictionary(wb):
    if "Data Dictionary" not in wb.sheetnames:
        return {}

    ws = wb["Data Dictionary"]
    dictionary = {}

    for row in ws.iter_rows(min_row=3, values_only=True):
        row  = list(row)
        prop = _cell(row, 0)
        if not prop or prop.startswith("─"):
            continue
        dictionary[prop] = {
            "allowed_values": _parse_allowed(_cell(row, 1)) if _cell(row, 1) else [],
            "type":           _normalize_type(_cell(row, 2)),
            "notes":          _cell(row, 3) or "",
        }

    return dictionary


def parse_tracking_plan(filepath) -> dict:
    """Main entry point. Returns the full schema dict."""
    wb = openpyxl.load_workbook(filepath, data_only=True)

    events          = parse_event_sheets(wb)
    global_props    = parse_global_props(wb)
    data_dictionary = parse_data_dictionary(wb)

    event_lookup = {
        ev["event_name"].lower().replace(" ", "_"): ev
        for ev in events
    }

    return {
        "events":          events,
        "event_lookup":    event_lookup,
        "global_props":    global_props,
        "data_dictionary": data_dictionary,
        "meta": {
            "source_file":   str(filepath),
            "total_events":  len(events),
            "sheets_parsed": EVENT_SHEETS,
        },
    }


def sample_events(events: list, config: dict) -> list:
    """
    Priority-based, session-preserving sampler.
    Never cuts a session in half — if one event from session X is selected,
    all events from that session are included.

    FIX 6 — sort uses _extract_event_time fallback so Export API events
    (which have event_time strings, not time ints) are ordered correctly.
    """
    from mcp_tools import _extract_event_time  # local import avoids circular dep at module load

    def _sort_key(e):
        t = e.get("time")
        if t and isinstance(t, (int, float)) and t > 0:
            return t
        dt = _extract_event_time(e)
        return int(dt.timestamp() * 1000) if dt else 0

    sample_size = config.get("sample_size", 500)

    if len(events) <= sample_size:
        return sorted(events, key=_sort_key)

    events_sorted = sorted(events, key=_sort_key)

    session_map = {}
    for ev in events_sorted:
        sid = (ev.get("event_properties") or {}).get("session_id") or f"u_{ev.get('user_id','anon')}"
        session_map.setdefault(sid, []).append(ev)

    with_orders   = [s for s in session_map.values() if any(e.get("event_type") == "Order Completed"  for e in s)]
    with_checkout = [s for s in session_map.values() if any(e.get("event_type") == "Checkout Started" for e in s)
                     and not any(e.get("event_type") == "Order Completed" for e in s)]
    others        = [s for s in session_map.values()
                     if not any(e.get("event_type") in ("Order Completed", "Checkout Started") for e in s)]

    sampled = []
    for bucket in (with_orders, with_checkout, others):
        for sess in bucket:
            if len(sampled) + len(sess) <= sample_size:
                sampled.extend(sess)

    return sorted(sampled, key=lambda e: e.get("time", 0))


if __name__ == "__main__":
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else "tracking_plan.xlsx"
    schema = parse_tracking_plan(path)
    print(json.dumps(schema, indent=2))
    print(f"\nParsed {schema['meta']['total_events']} events from {len(schema['meta']['sheets_parsed'])} sheets")
    # Show property counts per event for verification
    for ev in schema["events"]:
        print(f"  {ev['event_name']}: {len(ev['properties'])} properties")
